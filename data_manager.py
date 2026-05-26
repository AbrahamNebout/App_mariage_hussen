import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

EXCEL_FILE = "mariage_hussen_fransesca.xlsx"

SHEETS = {
    "Config": "Configuration",
    "Equipe": "Équipe & Tâches",
    "Deroulement": "Déroulement",
    "Checklist": "Checklist",
    "Budget": "Budget",
    "Invites": "Invités",
    "Fournisseurs": "Fournisseurs",
    "Notes": "Notes PCO",
}

GOLD = "C9A84C"
ROSE = "F4C2C2"
BLANC = "FFFDF5"
GRIS = "F2F2F2"
VERT = "90EE90"
ROUGE = "FFB6B6"

def get_border():
    thin = Side(style='thin', color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, bg=GOLD, fg="FFFFFF", size=11):
    cell.font = Font(bold=True, color=fg, size=size, name="Arial")
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = get_border()

def style_cell(cell, bg=BLANC):
    cell.font = Font(name="Arial", size=10)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = get_border()

def init_excel():
    if os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── CONFIG ──────────────────────────────────────────────────────────────
    ws = wb.create_sheet(SHEETS["Config"])
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 40
    config_data = [
        ("Marié", "Hussen Junior"),
        ("Mariée", "Fransesca"),
        ("Date du mariage", "01/08/2025"),
        ("Thème couleurs", "Bordeaux & Champagne Doré"),
        ("Couleur principale", "Bordeaux #800020"),
        ("Couleur secondaire", "Champagne Doré #C9A84C"),
        ("Couleur accent", "Ivoire #FFFDF5"),
        ("Lieu cérémonie", ""),
        ("Lieu réception", ""),
        ("PCO", ""),
        ("Nombre invités total", ""),
        ("Budget total estimé", ""),
        ("Budget dépensé", ""),
    ]
    ws["A1"] = "🌹 CONFIGURATION DU MARIAGE — Hussen & Fransesca"
    ws.merge_cells("A1:B1")
    style_header(ws["A1"], bg="800020", size=13)
    ws.row_dimensions[1].height = 35
    for i, (k, v) in enumerate(config_data, start=2):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
        style_cell(ws[f"A{i}"], bg=GRIS)
        style_cell(ws[f"B{i}"], bg=BLANC)

    # ── EQUIPE ──────────────────────────────────────────────────────────────
    ws2 = wb.create_sheet(SHEETS["Equipe"])
    headers = ["Nom", "Rôle / Fonction", "Tâches Assignées", "Téléphone", "Email", "Statut", "Notes"]
    widths = [22, 22, 45, 18, 28, 14, 30]
    ws2["A1"] = "👥 ÉQUIPE & TÂCHES — Mariage Hussen & Fransesca"
    ws2.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    style_header(ws2["A1"], bg="800020", size=13)
    ws2.row_dimensions[1].height = 35
    for j, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws2.cell(row=2, column=j, value=h)
        style_header(c, bg=GOLD)
        ws2.column_dimensions[get_column_letter(j)].width = w
    sample = [
        ["Hussen Junior", "Marié", "Coordination générale", "", "", "Actif", ""],
        ["Fransesca", "Mariée", "Coordination générale", "", "", "Actif", ""],
        ["", "Témoin du marié", "", "", "", "À confirmer", ""],
        ["", "Témoin de la mariée", "", "", "", "À confirmer", ""],
        ["", "MC / Animateur", "Animation réception", "", "", "À confirmer", ""],
        ["", "Responsable photo", "Photos & vidéo", "", "", "À confirmer", ""],
        ["", "Responsable accueil", "Accueil invités", "", "", "À confirmer", ""],
        ["", "Responsable transport", "Organisation navettes", "", "", "À confirmer", ""],
        ["", "Responsable décoration", "Mise en place déco", "", "", "À confirmer", ""],
        ["", "Responsable coordination salle", "Gestion salle réception", "", "", "À confirmer", ""],
    ]
    for r, row in enumerate(sample, start=3):
        for j, val in enumerate(row, start=1):
            c = ws2.cell(row=r, column=j, value=val)
            bg = GRIS if r % 2 == 0 else BLANC
            style_cell(c, bg=bg)

    # ── DEROULEMENT ─────────────────────────────────────────────────────────
    ws3 = wb.create_sheet(SHEETS["Deroulement"])
    headers3 = ["Étape", "Description", "Heure Prévue", "Lieu", "Responsable", "Tenue requise", "Photos prévues", "Statut", "Notes"]
    widths3 = [25, 40, 14, 25, 20, 25, 18, 14, 30]
    ws3["A1"] = "📋 DÉROULEMENT DU MARIAGE — Programme Détaillé"
    ws3.merge_cells(f"A1:{get_column_letter(len(headers3))}1")
    style_header(ws3["A1"], bg="800020", size=13)
    ws3.row_dimensions[1].height = 35
    for j, (h, w) in enumerate(zip(headers3, widths3), start=1):
        c = ws3.cell(row=2, column=j, value=h)
        style_header(c, bg=GOLD)
        ws3.column_dimensions[get_column_letter(j)].width = w
    programme = [
        ["Préparatifs Mariée", "Coiffure, maquillage, habillage mariée", "07:00", "Domicile mariée", "Témoin mariée", "Robe de préparation / kimono", "Oui - getting ready", "À planifier", ""],
        ["Préparatifs Marié", "Habillage et préparatifs du marié", "08:00", "Domicile marié", "Témoin marié", "Costume de cérémonie", "Oui - getting ready", "À planifier", ""],
        ["Photos avant cérémonie", "Séance photo couple avant église", "09:30", "Lieu à définir", "Photographe", "Tenue cérémonie complète", "Oui - séance principale", "À planifier", "First look optionnel"],
        ["Cortège vers l'église", "Départ en cortège vers l'église", "10:30", "Vers l'église", "Responsable transport", "Tenue cérémonie", "Oui - arrivée église", "À planifier", ""],
        ["Cérémonie religieuse", "Messe / cérémonie à l'église", "11:00", "Église", "Officiant", "Tenue cérémonie", "Oui - cérémonie", "À planifier", "Durée estimée: 1h30"],
        ["Sortie de l'église", "Riz / pétales / bulles de savon", "12:30", "Parvis église", "Responsable accueil", "Tenue cérémonie", "Oui - sortie église", "À planifier", "Prévoir confettis/pétales"],
        ["Photos officielles", "Photos avec famille et proches", "12:45", "Parvis / Jardin", "Photographe", "Tenue cérémonie", "Oui - photos famille", "À planifier", "Liste photos à préparer"],
        ["Trajet réception", "Déplacement vers salle de réception", "14:00", "En route", "Responsable transport", "Tenue cérémonie ou changement", "Non", "À planifier", ""],
        ["Cocktail / Vin d'honneur", "Accueil des invités, cocktail", "14:30", "Salle de réception", "MC + Responsable accueil", "Tenue cocktail (si changement)", "Oui - cocktail", "À planifier", ""],
        ["Entrée des mariés réception", "Entrée officielle dans la salle", "16:00", "Salle de réception", "MC", "Tenue soirée", "Oui - entrée salle", "À planifier", "Chanson d'entrée à choisir"],
        ["Discours & témoignages", "Discours témoins, familles", "16:30", "Salle de réception", "MC", "Tenue soirée", "Non", "À planifier", ""],
        ["Dîner de réception", "Repas assis ou buffet", "17:30", "Salle de réception", "Traiteur + MC", "Tenue soirée", "Oui - moments repas", "À planifier", ""],
        ["Gâteau de mariage", "Coupure du gâteau", "20:00", "Salle de réception", "MC + Traiteur", "Tenue soirée", "Oui - gâteau", "À planifier", ""],
        ["Première danse", "Ouverture du bal - première danse", "20:30", "Salle de réception", "MC + DJ", "Tenue soirée", "Oui", "À planifier", "Chanson à choisir"],
        ["Soirée dansante", "Fête, musique, danse", "21:00", "Salle de réception", "DJ / Animateur", "Tenue soirée", "Non", "À planifier", ""],
        ["Clôture & départ", "Fin de la réception", "02:00", "Salle de réception", "PCO + Responsable salle", "Tenue soirée", "Oui - feu d'artifice si prévu", "À planifier", ""],
    ]
    for r, row in enumerate(programme, start=3):
        for j, val in enumerate(row, start=1):
            c = ws3.cell(row=r, column=j, value=val)
            bg = GRIS if r % 2 == 0 else BLANC
            style_cell(c, bg=bg)

    # ── CHECKLIST ────────────────────────────────────────────────────────────
    ws4 = wb.create_sheet(SHEETS["Checklist"])
    headers4 = ["Catégorie", "Élément", "Détails / Description", "Lieu / Fournisseur", "Contact", "Prix Estimé (FCFA)", "Statut", "Date Confirmation", "Notes"]
    widths4 = [20, 28, 38, 25, 20, 18, 16, 18, 30]
    ws4["A1"] = "✅ CHECKLIST COMPLÈTE — Tout ce qu'il faut valider"
    ws4.merge_cells(f"A1:{get_column_letter(len(headers4))}1")
    style_header(ws4["A1"], bg="800020", size=13)
    ws4.row_dimensions[1].height = 35
    for j, (h, w) in enumerate(zip(headers4, widths4), start=1):
        c = ws4.cell(row=2, column=j, value=h)
        style_header(c, bg=GOLD)
        ws4.column_dimensions[get_column_letter(j)].width = w

    checklist_items = [
        # LIEU
        ("📍 Lieu", "Salle de réception", "Capacité, parking, accessibilité", "", "", "", "À confirmer", "", ""),
        ("📍 Lieu", "Église / Lieu cérémonie", "Réservation officiant", "", "", "", "À confirmer", "", ""),
        ("📍 Lieu", "Lieu séance photo", "Before/After cérémonie", "", "", "", "À confirmer", "", ""),
        ("📍 Lieu", "Hébergement nuit de noces", "Hôtel ou suite nuptiale", "", "", "", "À confirmer", "", ""),
        # TRAITEUR
        ("🍽️ Traiteur", "Service traiteur principal", "Menu, nombre couverts, service", "", "", "", "À confirmer", "", ""),
        ("🍽️ Traiteur", "Menu entrées", "Détail des entrées", "", "", "", "À confirmer", "", ""),
        ("🍽️ Traiteur", "Menu plats principaux", "Viande, poisson, végétarien", "", "", "", "À confirmer", "", ""),
        ("🍽️ Traiteur", "Menu desserts", "Hors gâteau de mariage", "", "", "", "À confirmer", "", ""),
        ("🍽️ Traiteur", "Cocktail / Vin d'honneur", "Canapés, finger food", "", "", "", "À confirmer", "", ""),
        # BOISSONS
        ("🥂 Boissons", "Champagne / Mousseux", "Pour toast et coupes", "", "", "", "À confirmer", "", ""),
        ("🥂 Boissons", "Vins (rouge, blanc, rosé)", "Accord mets-vins", "", "", "", "À confirmer", "", ""),
        ("🥂 Boissons", "Softs & jus de fruits", "Pour non-alcool", "", "", "", "À confirmer", "", ""),
        ("🥂 Boissons", "Eau (plate et gazeuse)", "", "", "", "", "À confirmer", "", ""),
        ("🥂 Boissons", "Bière et alcools divers", "", "", "", "", "À confirmer", "", ""),
        # GÂTEAU
        ("🎂 Gâteau", "Gâteau de mariage", "Design, saveurs, étages", "", "", "", "À confirmer", "", ""),
        ("🎂 Gâteau", "Gâteaux supplémentaires", "Pour le buffet desserts", "", "", "", "À confirmer", "", ""),
        # DÉCORATION
        ("🌸 Décoration", "Décoration salle réception", "Tables, centre de table, arche", "", "", "", "À confirmer", "", ""),
        ("🌸 Décoration", "Arche florale / cérémonie", "Fond de scène", "", "", "", "À confirmer", "", ""),
        ("🌸 Décoration", "Bouquet de la mariée", "Fleurs, style, couleurs", "", "", "", "À confirmer", "", ""),
        ("🌸 Décoration", "Boutonnieres & corsages", "Pour marié, témoins, parents", "", "", "", "À confirmer", "", ""),
        ("🌸 Décoration", "Centres de tables", "Fleurs ou décoration", "", "", "", "À confirmer", "", ""),
        ("🌸 Décoration", "Chemin de table", "", "", "", "", "À confirmer", "", ""),
        ("🌸 Décoration", "Bougies & éclairage déco", "", "", "", "", "À confirmer", "", ""),
        ("🌸 Décoration", "Ballons & guirlandes", "", "", "", "", "À confirmer", "", ""),
        ("🌸 Décoration", "Photobooth / Mur floral", "Installation photo", "", "", "", "À confirmer", "", ""),
        # TENUES
        ("👗 Tenues", "Robe de mariée", "Créateur, essayages", "", "", "", "À confirmer", "", ""),
        ("👗 Tenues", "Voile de la mariée", "", "", "", "", "À confirmer", "", ""),
        ("👗 Tenues", "Chaussures mariée", "Cérémonie + soirée", "", "", "", "À confirmer", "", ""),
        ("👗 Tenues", "Bijoux mariée", "Collier, boucles, bracelet", "", "", "", "À confirmer", "", ""),
        ("👗 Tenues", "Costume du marié", "Couleur, coupe, accessoires", "", "", "", "À confirmer", "", ""),
        ("👗 Tenues", "Chemise & cravate / nœud pap marié", "", "", "", "", "À confirmer", "", ""),
        ("👗 Tenues", "Chaussures marié", "", "", "", "", "À confirmer", "", ""),
        ("👗 Tenues", "Tenues demoiselles d'honneur", "Couleur coordonnée", "", "", "", "À confirmer", "", ""),
        ("👗 Tenues", "Tenues garçons d'honneur", "", "", "", "", "À confirmer", "", ""),
        ("👗 Tenues", "Tenue 2ème (soirée) mariée", "Changement pour réception", "", "", "", "À confirmer", "", ""),
        ("👗 Tenues", "Tenue 2ème (soirée) marié", "", "", "", "", "À confirmer", "", ""),
        # BEAUTÉ
        ("💄 Beauté", "Coiffure mariée", "Essai + jour J", "", "", "", "À confirmer", "", ""),
        ("💄 Beauté", "Maquillage mariée", "Essai + jour J", "", "", "", "À confirmer", "", ""),
        ("💄 Beauté", "Ongles mariée", "Manucure/pédicure", "", "", "", "À confirmer", "", ""),
        ("💄 Beauté", "Coiffure demoiselles d'honneur", "", "", "", "", "À confirmer", "", ""),
        # PHOTO / VIDÉO
        ("📸 Photo/Vidéo", "Photographe principal", "Package, heures, livrables", "", "", "", "À confirmer", "", ""),
        ("📸 Photo/Vidéo", "Vidéaste / Cinéaste", "Film de mariage", "", "", "", "À confirmer", "", ""),
        ("📸 Photo/Vidéo", "Drone", "Prises aériennes", "", "", "", "À confirmer", "", ""),
        ("📸 Photo/Vidéo", "Photobooth", "Machine + accessoires", "", "", "", "À confirmer", "", ""),
        # MUSIQUE
        ("🎵 Musique", "DJ / Sound system", "Soirée réception", "", "", "", "À confirmer", "", ""),
        ("🎵 Musique", "Groupe live / Chorale", "Cérémonie ou cocktail", "", "", "", "À confirmer", "", ""),
        ("🎵 Musique", "Sono cérémonie", "Micro officiant, musique", "", "", "", "À confirmer", "", ""),
        ("🎵 Musique", "Playlist sélectionnée", "Entrée, dîner, danse", "", "", "", "À confirmer", "", ""),
        # INVITATIONS
        ("💌 Papeterie", "Save the date", "Envoi anticipé", "", "", "", "À confirmer", "", ""),
        ("💌 Papeterie", "Cartes d'invitation", "Design, impression, envoi", "", "", "", "À confirmer", "", ""),
        ("💌 Papeterie", "Programme de cérémonie", "Livret messe / laïque", "", "", "", "À confirmer", "", ""),
        ("💌 Papeterie", "Menus de table", "Cartes menus imprimées", "", "", "", "À confirmer", "", ""),
        ("💌 Papeterie", "Plan de table", "Affichage + cartes de table", "", "", "", "À confirmer", "", ""),
        ("💌 Papeterie", "Livret d'accueil invités", "Infos pratiques", "", "", "", "À confirmer", "", ""),
        ("💌 Papeterie", "Étiquettes cadeaux/faveurs", "", "", "", "", "À confirmer", "", ""),
        # AFFICHES & SIGNALÉTIQUE
        ("🖼️ Affiches", "Affiche bienvenue entrée salle", "Avec photo du couple", "", "", "", "À confirmer", "", ""),
        ("🖼️ Affiches", "Panneau plan de table", "Grand format", "", "", "", "À confirmer", "", ""),
        ("🖼️ Affiches", "Enseigne chevalets directionnels", "Parking, salle, toilettes", "", "", "", "À confirmer", "", ""),
        ("🖼️ Affiches", "Tableau hashtag réseaux sociaux", "#HussenEtFransesca", "", "", "", "À confirmer", "", ""),
        ("🖼️ Affiches", "Affiches thématiques décoration", "Devis & photos de fiançailles", "", "", "", "À confirmer", "", ""),
        # TRANSPORT
        ("🚗 Transport", "Voiture des mariés", "Décoration, chauffeur", "", "", "", "À confirmer", "", ""),
        ("🚗 Transport", "Navettes invités", "Hôtel ↔ Salle", "", "", "", "À confirmer", "", ""),
        ("🚗 Transport", "Cortège voitures", "Organisation convoi", "", "", "", "À confirmer", "", ""),
        # HÉBERGEMENT
        ("🏨 Hébergement", "Suite nuptiale", "Nuit de noces", "", "", "", "À confirmer", "", ""),
        ("🏨 Hébergement", "Hébergements invités VIP", "Famille proche", "", "", "", "À confirmer", "", ""),
        # CADEAUX
        ("🎁 Cadeaux", "Urne ou liste de mariage", "", "", "", "", "À confirmer", "", ""),
        ("🎁 Cadeaux", "Faveurs / cadeaux invités", "Dragées, petits présents", "", "", "", "À confirmer", "", ""),
        ("🎁 Cadeaux", "Cadeaux témoins", "", "", "", "", "À confirmer", "", ""),
        # ADMINISTRATIF
        ("📄 Administratif", "Acte de mariage civil", "Mairie", "", "", "", "À confirmer", "", ""),
        ("📄 Administratif", "Livret de famille", "", "", "", "", "À confirmer", "", ""),
        ("📄 Administratif", "Assurance événement", "", "", "", "", "À confirmer", "", ""),
        # DIVERS
        ("🎆 Divers", "Feu d'artifice / Pétards", "Fin de soirée", "", "", "", "À confirmer", "", ""),
        ("🎆 Divers", "Lâcher de ballons / lanternes", "", "", "", "", "À confirmer", "", ""),
        ("🎆 Divers", "Livre d'or", "Avec polaroid ou signatures", "", "", "", "À confirmer", "", ""),
        ("🎆 Divers", "Candy bar / Buffet sucré", "", "", "", "", "À confirmer", "", ""),
        ("🎆 Divers", "Trône / Espace VIP mariés", "Chaises décorées, podium", "", "", "", "À confirmer", "", ""),
    ]
    cat_colors = {
        "📍 Lieu": "E8D5F5",
        "🍽️ Traiteur": "FFE4B5",
        "🥂 Boissons": "E0F7FA",
        "🎂 Gâteau": "FFD0D0",
        "🌸 Décoration": "F0FFF0",
        "👗 Tenues": "FFF0F5",
        "💄 Beauté": "FFEEF8",
        "📸 Photo/Vidéo": "E8F4FD",
        "🎵 Musique": "FFF8E1",
        "💌 Papeterie": "F5F5DC",
        "🖼️ Affiches": "E6E6FA",
        "🚗 Transport": "F0F8FF",
        "🏨 Hébergement": "F5FFFA",
        "🎁 Cadeaux": "FFF0E6",
        "📄 Administratif": "F0F0F0",
        "🎆 Divers": "FFFFF0",
    }
    for r, row in enumerate(checklist_items, start=3):
        cat = row[0]
        bg = cat_colors.get(cat, BLANC)
        for j, val in enumerate(row, start=1):
            c = ws4.cell(row=r, column=j, value=val)
            style_cell(c, bg=bg)

    # ── BUDGET ────────────────────────────────────────────────────────────
    ws5 = wb.create_sheet(SHEETS["Budget"])
    headers5 = ["Catégorie", "Poste de dépense", "Budget Estimé (FCFA)", "Acompte versé", "Reste à payer", "Payé ?", "Fournisseur", "Notes"]
    widths5 = [22, 30, 22, 20, 20, 12, 25, 30]
    ws5["A1"] = "💰 BUDGET — Suivi des Dépenses"
    ws5.merge_cells(f"A1:{get_column_letter(len(headers5))}1")
    style_header(ws5["A1"], bg="800020", size=13)
    ws5.row_dimensions[1].height = 35
    for j, (h, w) in enumerate(zip(headers5, widths5), start=1):
        c = ws5.cell(row=2, column=j, value=h)
        style_header(c, bg=GOLD)
        ws5.column_dimensions[get_column_letter(j)].width = w
    budget_items = [
        ("Lieu", "Salle de réception", 0, 0, "=C3-D3", "Non", "", ""),
        ("Lieu", "Église / offrande", 0, 0, "=C4-D4", "Non", "", ""),
        ("Traiteur", "Service traiteur complet", 0, 0, "=C5-D5", "Non", "", ""),
        ("Boissons", "Toutes boissons", 0, 0, "=C6-D6", "Non", "", ""),
        ("Gâteau", "Gâteau de mariage", 0, 0, "=C7-D7", "Non", "", ""),
        ("Décoration", "Fleurs + décoration salle", 0, 0, "=C8-D8", "Non", "", ""),
        ("Tenues", "Robe + costume + accessoires", 0, 0, "=C9-D9", "Non", "", ""),
        ("Beauté", "Coiffure + maquillage", 0, 0, "=C10-D10", "Non", "", ""),
        ("Photo/Vidéo", "Photographe + vidéaste", 0, 0, "=C11-D11", "Non", "", ""),
        ("Musique", "DJ + sono", 0, 0, "=C12-D12", "Non", "", ""),
        ("Papeterie", "Invitations + affiches", 0, 0, "=C13-D13", "Non", "", ""),
        ("Transport", "Voiture mariés + navettes", 0, 0, "=C14-D14", "Non", "", ""),
        ("Hébergement", "Suite nuptiale", 0, 0, "=C15-D15", "Non", "", ""),
        ("Cadeaux", "Faveurs + témoins", 0, 0, "=C16-D16", "Non", "", ""),
        ("Divers", "Imprévus (10%)", 0, 0, "=C17-D17", "Non", "", ""),
    ]
    for r, row in enumerate(budget_items, start=3):
        for j, val in enumerate(row, start=1):
            c = ws5.cell(row=r, column=j, value=val)
            bg = GRIS if r % 2 == 0 else BLANC
            style_cell(c, bg=bg)
    # Total row
    last = 3 + len(budget_items)
    ws5.cell(row=last, column=1, value="TOTAL").font = Font(bold=True, name="Arial")
    ws5.cell(row=last, column=3, value=f"=SUM(C3:C{last-1})")
    ws5.cell(row=last, column=4, value=f"=SUM(D3:D{last-1})")
    ws5.cell(row=last, column=5, value=f"=SUM(E3:E{last-1})")
    for j in range(1, 9):
        c = ws5.cell(row=last, column=j)
        style_header(c, bg=GOLD)

    # ── INVITÉS ────────────────────────────────────────────────────────────
    ws6 = wb.create_sheet(SHEETS["Invites"])
    headers6 = ["N°", "Nom & Prénom", "Côté (Marié/Mariée)", "Téléphone", "Email", "Table N°", "Menu choisi", "Allergie", "RSVP", "Transport prévu", "Hébergement", "Notes"]
    widths6 = [6, 25, 16, 18, 28, 10, 16, 18, 12, 18, 14, 25]
    ws6["A1"] = "👥 LISTE DES INVITÉS"
    ws6.merge_cells(f"A1:{get_column_letter(len(headers6))}1")
    style_header(ws6["A1"], bg="800020", size=13)
    ws6.row_dimensions[1].height = 35
    for j, (h, w) in enumerate(zip(headers6, widths6), start=1):
        c = ws6.cell(row=2, column=j, value=h)
        style_header(c, bg=GOLD)
        ws6.column_dimensions[get_column_letter(j)].width = w

    # ── FOURNISSEURS ────────────────────────────────────────────────────────
    ws7 = wb.create_sheet(SHEETS["Fournisseurs"])
    headers7 = ["Catégorie", "Nom société/Personne", "Service fourni", "Téléphone", "Email", "Adresse", "Montant contrat", "Acompte", "Reste", "Contrat signé ?", "Notes"]
    widths7 = [20, 25, 30, 18, 28, 30, 18, 15, 15, 15, 30]
    ws7["A1"] = "🤝 FOURNISSEURS & PRESTATAIRES"
    ws7.merge_cells(f"A1:{get_column_letter(len(headers7))}1")
    style_header(ws7["A1"], bg="800020", size=13)
    ws7.row_dimensions[1].height = 35
    for j, (h, w) in enumerate(zip(headers7, widths7), start=1):
        c = ws7.cell(row=2, column=j, value=h)
        style_header(c, bg=GOLD)
        ws7.column_dimensions[get_column_letter(j)].width = w
    prestataires = [
        ("Traiteur", "", "Repas complet", "", "", "", 0, 0, "=G3-H3", "Non", ""),
        ("Photo/Vidéo", "", "Photographe", "", "", "", 0, 0, "=G4-H4", "Non", ""),
        ("Photo/Vidéo", "", "Vidéaste", "", "", "", 0, 0, "=G5-H5", "Non", ""),
        ("Musique", "", "DJ", "", "", "", 0, 0, "=G6-H6", "Non", ""),
        ("Décoration", "", "Fleuriste", "", "", "", 0, 0, "=G7-H7", "Non", ""),
        ("Beauté", "", "Coiffeuse mariée", "", "", "", 0, 0, "=G8-H8", "Non", ""),
        ("Beauté", "", "Maquilleuse mariée", "", "", "", 0, 0, "=G9-H9", "Non", ""),
        ("Gâteau", "", "Pâtissier", "", "", "", 0, 0, "=G10-H10", "Non", ""),
        ("Transport", "", "Chauffeur voiture mariés", "", "", "", 0, 0, "=G11-H11", "Non", ""),
        ("Papeterie", "", "Imprimeur invitations", "", "", "", 0, 0, "=G12-H12", "Non", ""),
    ]
    for r, row in enumerate(prestataires, start=3):
        for j, val in enumerate(row, start=1):
            c = ws7.cell(row=r, column=j, value=val)
            bg = GRIS if r % 2 == 0 else BLANC
            style_cell(c, bg=bg)

    # ── NOTES PCO ────────────────────────────────────────────────────────────
    ws8 = wb.create_sheet(SHEETS["Notes"])
    headers8 = ["Date", "Sujet", "Note / Décision prise", "Action requise", "Responsable", "Statut", "Priorité"]
    widths8 = [14, 25, 50, 35, 20, 14, 12]
    ws8["A1"] = "📝 JOURNAL PCO — Notes & Décisions"
    ws8.merge_cells(f"A1:{get_column_letter(len(headers8))}1")
    style_header(ws8["A1"], bg="800020", size=13)
    ws8.row_dimensions[1].height = 35
    for j, (h, w) in enumerate(zip(headers8, widths8), start=1):
        c = ws8.cell(row=2, column=j, value=h)
        style_header(c, bg=GOLD)
        ws8.column_dimensions[get_column_letter(j)].width = w
    ws8.cell(row=3, column=1, value=datetime.now().strftime("%d/%m/%Y"))
    ws8.cell(row=3, column=2, value="Initialisation")
    ws8.cell(row=3, column=3, value="Fichier PCO créé — Mariage Hussen Junior & Fransesca — 01/08/2025")
    ws8.cell(row=3, column=4, value="Commencer à remplir les informations")
    ws8.cell(row=3, column=5, value="PCO")
    ws8.cell(row=3, column=6, value="Fait")
    ws8.cell(row=3, column=7, value="🔴 Urgent")
    for j in range(1, 8):
        style_cell(ws8.cell(row=3, column=j), bg=BLANC)

    wb.save(EXCEL_FILE)
    print(f"✅ Fichier Excel créé: {EXCEL_FILE}")

def load_sheet(sheet_name):
    if not os.path.exists(EXCEL_FILE):
        init_excel()
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name, header=1)
        return df
    except Exception:
        return pd.DataFrame()

def save_sheet(sheet_name, df):
    if not os.path.exists(EXCEL_FILE):
        init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    # Find header row and data start
    max_row = ws.max_row
    # Remove old data rows (keep row 1 = titre, row 2 = headers)
    for row in ws.iter_rows(min_row=3, max_row=max_row):
        for cell in row:
            cell.value = None
    # Write new data
    for r_idx, row in df.iterrows():
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx + 3, column=c_idx, value=val)
            bg = GRIS if (r_idx + 3) % 2 == 0 else BLANC
            style_cell(c, bg=bg)
    wb.save(EXCEL_FILE)

def get_config():
    if not os.path.exists(EXCEL_FILE):
        init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEETS["Config"]]
    config = {}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if row[0] and row[1] is not None:
            config[row[0]] = row[1]
    return config

def save_config(config_dict):
    if not os.path.exists(EXCEL_FILE):
        init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb[SHEETS["Config"]]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        key = row[0].value
        if key and key in config_dict:
            row[1].value = config_dict[key]
    wb.save(EXCEL_FILE)

def get_sheets_names():
    return SHEETS
